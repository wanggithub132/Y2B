from openpyxl import Workbook


# 将list写入excel
def write_to_excel(data_list, file_path):
    wb = Workbook()
    ws = wb.active

    # 写入表头
    header = list(data_list[0].keys())
    for i, col in enumerate(header):
        ws.cell(row=1, column=i + 1, value=col)

    # 写入数据
    for row_idx, data in enumerate(data_list, start=2):
        for col_idx, value in enumerate(data.values()):
            ws.cell(row=row_idx, column=col_idx + 1, value=value)

    wb.save(file_path)
